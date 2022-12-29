import pyautogui
import openpyxl
from datetime import datetime, timedelta

class Employee:
    def __init__(self, name, number, hours = timedelta(hours = 0), OThours = timedelta(hours = 0)):
        self.name = name
        self.number = number
        self.hours = hours
        self.OThours = OThours

def timeCalculator(tempCellValue, employeeList, employeeNumber):
    FMT = '%H:%M'
    OT = datetime.strptime("08:00:00", "%H:%M:%S")

    tempCellValue = tempCellValue.strip()
    timeSplitted = tempCellValue.splitlines()
    clockIn = timeSplitted[0]
    clockOut = timeSplitted[-1]
    print(clockIn)
    print(clockOut)
    tdelta = datetime.strptime(clockOut, FMT) - datetime.strptime(clockIn, FMT)

    time = str(tdelta)
    tempSplitted = time.split(", ")
    time = tempSplitted[-1]
    tdelta = datetime.strptime(time, "%H:%M:%S")

    if tdelta > OT:
        OThours = tdelta - timedelta(hours = 8)
        for i in employeeList:
            if employeeNumber == i.number:
                i.hours = i.hours + timedelta(hours = 8)
                i.OThours = i.OThours + OThours
       
    for x in employeeList:
        print(x.name)
        print(x.number)
        print(x.hours)
        print(x.OThours)

def readFile():
    employeeList = []
    try:
        x = pyautogui.prompt('Please enter excel timesheet filename', 'LostSG Salary Calculator')
        if x == "":
            pyautogui.alert(text='Please enter filename!', title='Error', button='OK')
        if x is None:
            pass
        else:
            x = x + ".xlsx"
            print(x)
            punchTime = openpyxl.load_workbook(x)
            punchTime.active = punchTime['Punch Time']
            sh = punchTime.active
            #for i in range(1, sh.max_row+1):
            for i in range(5,6):
                print("\n")
                print("Row ", i, " data :")
                for j in range(1, sh.max_column+1):
                    if j == 1:
                        cell_obj = sh.cell(row=i, column=j)
                        temp = Employee(cell_obj.value, i)
                        employeeList.append(temp)
                        print(temp.name)
                        print(temp.number)
                    cell_obj = sh.cell(row=i, column=j)
                    tempCellValue = cell_obj.value
                    if j > 1:
                        try:
                            if tempCellValue is None:
                                pass
                            else:
                                tempCellValue = tempCellValue.replace(" ","")
                                timeCalculator(tempCellValue, employeeList, temp.number)
                        except ValueError:
                            pass
                        except AttributeError:
                            pass

    except FileNotFoundError:
        pyautogui.alert(text='File not Found', title='Error', button='OK')

def main():
    readFile()
if __name__ == '__main__':
    main()